# import csv

# lines = list()
# memberName = input("Please enter a member's name to be deleted.")
# with open('Cytomat_inventory.csv', 'r') as readFile:
#     reader = csv.reader(readFile)
#     for row in reader:
#         lines.append(row)
#         for field in row:
#             if field == memberName:
#                 lines.remove(row)
# with open('Cytomat_inventory.csv', 'w') as writeFile:
#     writer = csv.writer(writeFile)
#     writer.writerows(lines)
from openpyxl import Workbook
from openpyxl import load_workbook

w = Workbook()
sheet = w.active
sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 30
sheet.column_dimensions['C'].width = 30
sheet.column_dimensions['D'].width = 30

sheet.cell(row=1,column=1).value = "Date"
sheet.cell(row=1,column=2).value = "Plate ID"
sheet.cell(row=1,column=3).value = "Location"
sheet.cell(row=1,column=4).value = "Researcher"

sheet.cell(row=2,column=1).value = "x"
sheet.cell(row=2,column=2).value = "y ID"
sheet.cell(row=2,column=3).value = "z"
sheet.cell(row=2,column=4).value = "r"

sheet.cell(row=3,column=1).value = "xa"
sheet.cell(row=3,column=2).value = "bb"
sheet.cell(row=3,column=3).value = "cc"
sheet.cell(row=3,column=4).value = "dd"

sheet.delete_rows(2)

w.save(filename="tests.xlsx")